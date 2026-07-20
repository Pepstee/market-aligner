"""
skeleton/reporter.py — the thin reporter the skeleton drives (Architecture.md
"Outputs"; Build-Spec §6).

Input: a list of C4 ScoredRow. Output, all under ``outputs/``:

    jobs_ranked.xlsx
        • sheet "jobs"         — one row per posting, sorted by Final desc,
                                 columns from §3 + Fit / Opportunity / Final.
        • sheet "field_ranked" — career, field_score, #jobs, #entry-level,
                                 median Fit  (the actual "which field" table).
    requirements_ranked.xlsx
        • canonical skill, frequency, % of postings, top fields demanding it —
          Artiom's evidence/portfolio to-do list.
    fit_opportunity.png
        • Fit-vs-Opportunity 2D scatter (high-fit/high-opp = sweet spot).

Deps: openpyxl (workbooks) + matplotlib (scatter). Both are optional at import
time — a caller can build just the workbooks on a box without matplotlib, and
each writer degrades to a clear error only if actually invoked without its dep.

    pip install openpyxl matplotlib --break-system-packages

Talks to the rest of the system only through skeleton/contracts.py (C4).
"""

from __future__ import annotations

import math
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional, Sequence

try:  # pragma: no cover - import shim
    from skeleton.contracts import ScoredRow
    from skeleton.scoring import aggregate_fields, FieldScore
except ImportError:  # pragma: no cover - import shim
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from contracts import ScoredRow  # type: ignore
    from scoring import aggregate_fields, FieldScore  # type: ignore


DEFAULT_OUTPUT_DIR = "outputs"

# Columns for the jobs sheet — provenance + §3 schema + scores.
JOB_COLUMNS: tuple[str, ...] = (
    "board", "job_id", "url", "posted_at", "scraped_at",
    "job_title", "company", "location", "salary_text", "contract_type",
    "experience_required", "sponsorship_signal", "mapped_career", "entry_level", "remote_flag",
    "job_description", "responsibilities", "required_software", "required_skills",
    "preferred_skills", "education_required", "certifications_required",
    "benefits", "application_deadline",
    "technical_alignment", "evidence_match", "growth_potential",
    "market_demand", "barrier_to_entry",
    "why_it_fits", "skills_to_learn",
    "extraction_confidence", "dedup_key",
    "fit", "opportunity", "final",
)


@dataclass
class ReportPaths:
    jobs_xlsx: Path
    requirements_xlsx: Path
    scatter_png: Path
    shortlist_md: Path


# --------------------------------------------------------------------------- #
# Public entry point.
# --------------------------------------------------------------------------- #
def write_reports(
    scored: Sequence[ScoredRow],
    output_dir: str | Path = DEFAULT_OUTPUT_DIR,
    make_plot: bool = True,
) -> ReportPaths:
    """Write both workbooks (+ the scatter) from a list of C4 rows.

    Returns the paths written. `make_plot=False` skips the PNG (useful when
    matplotlib is absent). The workbooks are always written.
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    ranked = sorted(scored, key=lambda sr: sr.final, reverse=True)
    fields = aggregate_fields(ranked)

    paths = ReportPaths(
        jobs_xlsx=out / "jobs_ranked.xlsx",
        requirements_xlsx=out / "requirements_ranked.xlsx",
        scatter_png=out / "fit_opportunity.png",
        shortlist_md=out / "SHORTLIST.md",
    )

    write_jobs_workbook(ranked, fields, paths.jobs_xlsx)
    write_requirements_workbook(ranked, paths.requirements_xlsx)
    write_shortlist(ranked, paths.shortlist_md)
    if make_plot:
        try:
            write_scatter(ranked, paths.scatter_png)
        except ImportError as e:  # matplotlib missing → skip, don't crash the run
            print(f"[reporter] scatter skipped ({e}); workbooks written.")
    return paths


def write_shortlist(
    ranked: Sequence[ScoredRow], path: str | Path, limit: int = 12
) -> Path:
    """Write an actionable shortlist locally with no model call."""
    eligible = [
        sr for sr in ranked
        if sr.row.mapped_career != "other"
        and (sr.row.technical_alignment or 0) >= 5
        and sr.final >= 65
    ]
    eligible.sort(key=lambda sr: (sr.row.entry_level is True, sr.final), reverse=True)
    eligible = eligible[:limit]
    lines = [
        "# UK application shortlist",
        "",
        "Vacancies were extracted and aligned by the configured LLM, then ranked by deterministic scoring. Scores are uncalibrated; verify work rights and closing dates before applying.",
        "",
    ]
    for index, sr in enumerate(eligible, 1):
        tier = "APPLY" if sr.row.entry_level is True else "REACH"
        lines.extend([
            f"## {index}. [{sr.row.job_title}]({sr.row.url}) — {tier}",
            "",
            f"- Employer/location: {sr.row.company} — {sr.row.location or 'not stated'}",
            f"- Track/score: {sr.row.mapped_career} — {sr.final:.1f}/100",
            f"- Experience: {sr.row.experience_required or 'not explicit'}",
            f"- Sponsorship: {sr.row.sponsorship_signal or 'unknown'}",
            f"- Fit note: {sr.row.why_it_fits}",
            "",
        ])
    path = Path(path)
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


# --------------------------------------------------------------------------- #
# jobs_ranked.xlsx  (jobs + field_ranked sheets)
# --------------------------------------------------------------------------- #
def write_jobs_workbook(
    ranked: Sequence[ScoredRow],
    fields: Sequence[FieldScore],
    path: str | Path,
) -> Path:
    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.active
    ws.title = "jobs"
    ws.append(["hiring_likelihood_rank", *JOB_COLUMNS])
    for rank, sr in enumerate(ranked, 1):
        ws.append([rank, *[_cell(_job_value(sr, col)) for col in JOB_COLUMNS]])

    fs = wb.create_sheet("field_ranked")
    fs.append(["career", "field_score", "n_jobs", "n_entry_level", "median_fit"])
    for f in fields:
        fs.append([
            f.career,
            round(f.field_score, 4),
            f.n_jobs,
            f.n_entry_level,
            round(f.median_top_fit, 4),
        ])

    _autosize(ws)
    _autosize(fs)
    _freeze_header(ws)
    _freeze_header(fs)

    path = Path(path)
    wb.save(path)
    return path


def _job_value(sr: ScoredRow, col: str) -> Any:
    if col == "fit":
        return round(sr.fit, 4)
    if col == "opportunity":
        return round(sr.opportunity, 4)
    if col == "final":
        return round(sr.final, 2)
    return getattr(sr.row, col, None)


# --------------------------------------------------------------------------- #
# requirements_ranked.xlsx  (canonical skill frequency = portfolio to-do list)
# --------------------------------------------------------------------------- #
def skill_frequency(ranked: Sequence[ScoredRow]) -> list[dict[str, Any]]:
    """Canonical-skill frequency table across all postings.

    +1 per canonical skill id per posting it appears in (Build-Spec §4), plus
    the % of postings demanding it and the top mapped_careers that demand it.
    """
    total = len(ranked)
    required: Counter[str] = Counter()
    preferred: Counter[str] = Counter()
    freq: Counter[str] = Counter()
    by_field: dict[str, Counter[str]] = defaultdict(Counter)

    for sr in ranked:
        career = sr.row.mapped_career or "other"
        # A skill counts once per posting even if listed twice.
        required_set = {s for s in ((sr.row.required_software or []) + (sr.row.required_skills or [])) if s}
        preferred_set = {s for s in (sr.row.preferred_skills or []) if s}
        for skill in required_set:
            required[skill] += 1
        for skill in preferred_set:
            preferred[skill] += 1
        for skill in required_set | preferred_set:
            freq[skill] += 1
            by_field[skill][career] += 1

    rows: list[dict[str, Any]] = []
    for skill, count in freq.most_common():
        top_fields = ", ".join(
            f"{career}({n})" for career, n in by_field[skill].most_common(3)
        )
        rows.append({
            "skill": skill,
            "required_frequency": required[skill],
            "preferred_frequency": preferred[skill],
            "any_frequency": count,
            "pct_of_postings": round(100.0 * count / total, 1) if total else 0.0,
            "top_fields": top_fields,
        })
    return rows


def write_requirements_workbook(ranked: Sequence[ScoredRow], path: str | Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "requirements"
    ws.append(["skill", "required_frequency", "preferred_frequency", "any_frequency", "pct_of_postings", "top_fields"])
    for row in skill_frequency(ranked):
        ws.append([row["skill"], row["required_frequency"], row["preferred_frequency"],
                   row["any_frequency"], row["pct_of_postings"], row["top_fields"]])

    _autosize(ws)
    _freeze_header(ws)

    path = Path(path)
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# fit_opportunity.png  (the 2D map that answers "which field")
# --------------------------------------------------------------------------- #
def write_scatter(ranked: Sequence[ScoredRow], path: str | Path) -> Path:
    import matplotlib

    matplotlib.use("Agg")  # headless — no display needed
    import matplotlib.pyplot as plt

    # One colour per career; postings placed at (Fit, Opportunity).
    careers = sorted({sr.row.mapped_career or "other" for sr in ranked})
    cmap = plt.get_cmap("tab10")
    colour = {c: cmap(i % 10) for i, c in enumerate(careers)}

    fig, ax = plt.subplots(figsize=(8, 6))
    for c in careers:
        pts = [(sr.fit, sr.opportunity) for sr in ranked if (sr.row.mapped_career or "other") == c]
        if not pts:
            continue
        xs, ys = zip(*pts)
        ax.scatter(xs, ys, label=c, color=colour[c], alpha=0.75, edgecolors="none", s=42)

    # Quadrant guides at the midpoint — sweet spot is top-right.
    ax.axvline(0.5, color="0.8", lw=1, zorder=0)
    ax.axhline(0.5, color="0.8", lw=1, zorder=0)
    ax.text(0.98, 0.98, "sweet spot", ha="right", va="top",
            transform=ax.transAxes, fontsize=9, color="0.4")
    ax.text(0.02, 0.98, "passion,\nthin market", ha="left", va="top",
            transform=ax.transAxes, fontsize=9, color="0.4")

    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.set_xlabel("Fit  (profile evidence × vacancy alignment)")
    ax.set_ylabel("Opportunity  (demand × accessibility × growth)")
    ax.set_title("Fit vs Opportunity — per posting")
    ax.legend(loc="lower left", fontsize=8, framealpha=0.9, ncol=2)
    fig.tight_layout()

    path = Path(path)
    fig.savefig(path, dpi=120)
    plt.close(fig)
    return path


# --------------------------------------------------------------------------- #
# small openpyxl helpers
# --------------------------------------------------------------------------- #
def _cell(value: Any) -> Any:
    """Coerce a Python value into something openpyxl can store in a cell."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (list, tuple)):
        return ", ".join(str(v) for v in value)
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        return value
    return value


def _autosize(ws, max_width: int = 60) -> None:
    from openpyxl.utils import get_column_letter

    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max_width, w + 2)


def _freeze_header(ws) -> None:
    ws.freeze_panes = "A2"


# --------------------------------------------------------------------------- #
# Self-test: `python skeleton/reporter.py` → writes to outputs/_reporter_selftest/
# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from contracts import JobRow  # type: ignore

    rows = []
    for i in range(6):
        rows.append(
            ScoredRow(
                row=JobRow(
                    board="wanted", job_id=str(i), url=f"https://x/{i}",
                    job_title=f"Job {i}", company=f"Co {i}",
                    mapped_career="UX_UI" if i % 2 == 0 else "ArchViz",
                    entry_level=(i % 3 == 0),
                    required_software=["figma", "blender"][: (i % 2) + 1],
                ),
                fit=0.4 + 0.08 * i,
                opportunity=0.9 - 0.1 * i,
                final=50.0 + i,
            )
        )
    paths = write_reports(rows, output_dir="outputs/_reporter_selftest")
    assert paths.jobs_xlsx.exists() and paths.requirements_xlsx.exists()
    print(f"reporter.py OK — wrote {paths.jobs_xlsx}, {paths.requirements_xlsx}, "
          f"{paths.scatter_png} (exists={paths.scatter_png.exists()})")
