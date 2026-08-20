"""
skeleton/test_skeleton.py — unit tests for the deterministic maths + reporter.

Run:  python3 skeleton/test_skeleton.py        (or under pytest)

Covers, per the build brief:
  Scoring (power/geometric mean §5):
    * all-equal inputs return that value           (mean is a true average)
    * one low factor drags the score down          (no over-forgiveness)
    * ε floor prevents zero-collapse               (one weak axis ≠ dead row)
    * monotonic in each axis                        (raise any axis → score ↑)
    * field aggregation + sensitivity smoke         (§5 decision maths)
  Reporter:
    * both xlsx build from a 20-row synthetic C4 fixture, with the right
      sheets/headers, sorted by Final.

Uses only a tiny synthetic fixture — no scraper, llm, or profiler needed.
"""

from __future__ import annotations

import math
import sys
from pathlib import Path
from types import SimpleNamespace

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE))

import scoring  # noqa: E402
import reporter  # noqa: E402
import run as pipeline_run  # noqa: E402
from contracts import (  # noqa: E402
    CandidateFitProfile,
    CandidateTrackProfile,
    JobRow,
    ScoredRow,
    JobUrl,
    RawPosting,
    write_jsonl,
)


# --------------------------------------------------------------------------- #
# Shared fixtures.
# --------------------------------------------------------------------------- #
def _params(p: float = 0.0, eps: float = 0.05) -> scoring.ScoringParams:
    return scoring.ScoringParams.from_config(
        {
            "scoring": {
                "mean_p": p,
                "epsilon": eps,
                "fit_weights": {
                    "interest": 0.20, "skill_alignment": 0.20,
                    "market_readiness": 0.20, "technical_alignment": 0.20,
                    "evidence_match": 0.20,
                },
                "opportunity_weights": {
                    "market_demand": 0.35, "accessibility": 0.35,
                    "growth_potential": 0.30,
                },
                "fit_opportunity_blend": 0.7,
            }
        }
    )


def _profile() -> CandidateFitProfile:
    return CandidateFitProfile(
        fields={
            "Agentic_AI_Engineer": CandidateTrackProfile(
                interest=10.0, skill=8.0, confidence=0.85, market_readiness=7.0
            ),
            "Backend_Engineer": CandidateTrackProfile(
                interest=7.5, skill=6.5, confidence=0.70, market_readiness=6.0
            ),
            "Full_Stack_Engineer": CandidateTrackProfile(
                interest=6.0, skill=5.0, confidence=0.55, market_readiness=5.0
            ),
        }
    )


def _equal(a: float, b: float, tol: float = 1e-9) -> bool:
    return abs(a - b) <= tol


# --------------------------------------------------------------------------- #
# power_mean — the core §5 maths.
# --------------------------------------------------------------------------- #
def test_all_equal_inputs_return_that_value():
    """A weighted geometric (and any power) mean of all-equal values = the value."""
    w = [0.25, 0.25, 0.25, 0.25]
    for v in (0.2, 0.5, 0.9):
        for p in (0.0, 1.0, -1.0, 2.0):
            got = scoring.power_mean([v, v, v, v], w, p=p, epsilon=0.01)
            assert _equal(got, v, 1e-9), f"all-equal {v} at p={p} → {got}"


def test_one_low_factor_drags_score():
    """Dropping one axis to near-zero must pull the mean well below the others."""
    w = [0.25, 0.25, 0.25, 0.25]
    high = scoring.power_mean([0.9, 0.9, 0.9, 0.9], w, p=0.0, epsilon=0.05)
    dragged = scoring.power_mean([0.9, 0.9, 0.9, 0.1], w, p=0.0, epsilon=0.05)
    assert dragged < high, "one low factor did not drag the geometric mean down"
    # Geometric mean punishes harder than the arithmetic mean would.
    arith = sum([0.9, 0.9, 0.9, 0.1]) / 4
    assert dragged < arith, "geometric mean should punish the low factor below arithmetic"


def test_epsilon_prevents_zero_collapse():
    """A zero (or missing) axis floors at ε, so the row survives instead of dying."""
    w = [0.25, 0.25, 0.25, 0.25]
    got = scoring.power_mean([0.9, 0.9, 0.9, 0.0], w, p=0.0, epsilon=0.05)
    assert got > 0.0, "ε floor failed — a single zero collapsed the whole mean"
    # With ε it equals the geo mean using ε in place of the zero.
    expected = math.exp(0.25 * (3 * math.log(0.9) + math.log(0.05)))
    assert _equal(got, expected, 1e-9)
    # Smaller ε → harsher penalty for the weak axis.
    harsher = scoring.power_mean([0.9, 0.9, 0.9, 0.0], w, p=0.0, epsilon=0.01)
    assert harsher < got, "smaller ε should penalise the zero axis more"


def test_monotonic_in_each_axis():
    """Raising any single axis (others fixed) must not lower the mean."""
    w = [0.25, 0.25, 0.25, 0.25]
    base = [0.4, 0.5, 0.6, 0.7]
    for p in (0.0, 1.0, -1.0):
        base_score = scoring.power_mean(base, w, p=p, epsilon=0.05)
        for i in range(4):
            bumped = list(base)
            bumped[i] = min(1.0, bumped[i] + 0.2)
            got = scoring.power_mean(bumped, w, p=p, epsilon=0.05)
            assert got >= base_score - 1e-12, f"non-monotone at axis {i}, p={p}"


def test_power_mean_ordering():
    """For non-equal inputs: harmonic ≤ geometric ≤ arithmetic (power-mean inequality)."""
    vals = [0.3, 0.6, 0.9]
    w = [1 / 3, 1 / 3, 1 / 3]
    harmonic = scoring.power_mean(vals, w, p=-1.0, epsilon=0.01)
    geometric = scoring.power_mean(vals, w, p=0.0, epsilon=0.01)
    arithmetic = scoring.power_mean(vals, w, p=1.0, epsilon=0.01)
    assert harmonic < geometric < arithmetic


# --------------------------------------------------------------------------- #
# score_row / axes — the join with candidate priors.
# --------------------------------------------------------------------------- #
def test_accessibility_inverts_barrier():
    assert _equal(scoring.accessibility(0.0), 1.0)
    assert _equal(scoring.accessibility(10.0), 0.0)
    assert _equal(scoring.accessibility(3.0), 0.7)


def test_score_row_uses_profile_and_bounds():
    params, prof = _params(), _profile()
    row = JobRow(
        board="greenhouse", job_id="1", url="u", mapped_career="Agentic_AI_Engineer",
        entry_level=True, technical_alignment=8.0, evidence_match=7.0,
        growth_potential=8.0, market_demand=9.0, barrier_to_entry=2.0,
    )
    sr = scoring.score_row(row, prof, params)
    assert 0.0 <= sr.fit <= 1.0 and 0.0 <= sr.opportunity <= 1.0
    assert 0.0 <= sr.final <= 100.0
    # A stronger profile track should out-Fit a weaker one,
    # holding posting axes constant.
    row2 = JobRow(**{**row.__dict__, "mapped_career": "Full_Stack_Engineer"})
    sr2 = scoring.score_row(row2, prof, params)
    assert sr.fit > sr2.fit, "stronger profile field should yield higher Fit"


def test_final_blend_formula():
    """final == 100·(blend·fit + (1-blend)·opp) exactly."""
    params, prof = _params(), _profile()
    row = JobRow(
        board="greenhouse", job_id="1", url="u", mapped_career="Backend_Engineer",
        entry_level=True, technical_alignment=6.0, evidence_match=6.0,
        growth_potential=7.0, market_demand=6.0, barrier_to_entry=4.0,
    )
    sr = scoring.score_row(row, prof, params)
    expected = 100.0 * (params.blend * sr.fit + (1 - params.blend) * sr.opportunity)
    assert _equal(sr.final, expected, 1e-9)


# --------------------------------------------------------------------------- #
# Field aggregation + sensitivity (§5 decision maths).
# --------------------------------------------------------------------------- #
def _synthetic_scored(n: int = 20) -> list[ScoredRow]:
    """A deterministic 20-row synthetic C4 fixture across three careers."""
    careers = ["Agentic_AI_Engineer", "Backend_Engineer", "Full_Stack_Engineer"]
    skills = [["agents", "python"], ["fastapi", "sql"], ["typescript", "react"], ["aws", "docker"]]
    rows: list[ScoredRow] = []
    for i in range(n):
        career = careers[i % len(careers)]
        row = JobRow(
            board="greenhouse", job_id=str(i), url=f"https://example.test/{i}",
            job_title=f"{career} role {i}", company=f"Company {i}",
            mapped_career=career, entry_level=(i % 2 == 0),
            required_software=skills[i % len(skills)],
            technical_alignment=float((i % 10)), evidence_match=float((i * 3) % 11),
            market_demand=float((i * 7) % 11), barrier_to_entry=float((i * 5) % 11),
            growth_potential=float((i * 2) % 11),
            extraction_confidence=0.8, dedup_key=f"company {i}|{career}",
        )
        # Fit/opp/final values exercise report sorting and aggregation.
        fit = 0.30 + 0.03 * (i % 12)
        opp = 0.80 - 0.02 * (i % 15)
        final = 100.0 * (0.6 * fit + 0.4 * opp)
        rows.append(ScoredRow(row=row, fit=fit, opportunity=opp, final=final))
    return rows


def test_field_aggregation_shape():
    scored = _synthetic_scored(20)
    fields = scoring.aggregate_fields(scored)
    assert fields, "expected at least one field"
    # Sorted descending by field_score.
    scores = [f.field_score for f in fields]
    assert scores == sorted(scores, reverse=True)
    # Counts add up to the total rows.
    assert sum(f.n_jobs for f in fields) == 20
    # field_score = median(top-10 fit) * log(1 + entry_count) — verify one field.
    ux = next(f for f in fields if f.career == "Agentic_AI_Engineer")
    ux_rows = [sr for sr in scored if sr.row.mapped_career == "Agentic_AI_Engineer"]
    import statistics as _st
    top = sorted((sr.fit for sr in ux_rows), reverse=True)[:10]
    entry = sum(1 for sr in ux_rows if sr.row.entry_level)
    assert _equal(ux.field_score, _st.median(top) * math.log(1 + entry), 1e-9)


def test_sensitivity_reports_stability():
    # Build C3 rows whose top field is obvious and robust: Agentic AI dominates.
    prof = _profile()
    params = _params()
    rows: list[JobRow] = []
    for i in range(15):
        career = (
            "Agentic_AI_Engineer" if i < 9
            else ("Backend_Engineer" if i < 12 else "Full_Stack_Engineer")
        )
        rows.append(JobRow(
            board="greenhouse", job_id=str(i), url="u", mapped_career=career,
            entry_level=True,
            technical_alignment=9.0 if career == "Agentic_AI_Engineer" else 3.0,
            evidence_match=9.0 if career == "Agentic_AI_Engineer" else 3.0,
            market_demand=8.0, barrier_to_entry=2.0, growth_potential=6.0,
        ))
    rep = scoring.sensitivity(rows, prof, params, delta=0.2)
    assert rep.baseline_top == "Agentic_AI_Engineer"
    assert rep.n_variants > 0
    assert rep.stable is True, rep.summary()


# --------------------------------------------------------------------------- #
# Reporter — build both workbooks from the 20-row synthetic C4 fixture.
# --------------------------------------------------------------------------- #
def test_reporter_builds_both_workbooks(tmp_dir: Path):
    from openpyxl import load_workbook

    scored = _synthetic_scored(20)
    paths = reporter.write_reports(scored, output_dir=tmp_dir, make_plot=True)

    assert paths.jobs_xlsx.exists(), "jobs_ranked.xlsx not written"
    assert paths.requirements_xlsx.exists(), "requirements_ranked.xlsx not written"

    # jobs workbook: two sheets, header matches, sorted by Final desc, 20 rows.
    wb = load_workbook(paths.jobs_xlsx)
    assert wb.sheetnames == ["jobs", "field_ranked"], wb.sheetnames
    jobs = wb["jobs"]
    header = [c.value for c in jobs[1]]
    assert header == ["hiring_likelihood_rank", *reporter.JOB_COLUMNS]
    finals = [row[1 + reporter.JOB_COLUMNS.index("final")].value for row in jobs.iter_rows(min_row=2)]
    assert len(finals) == 20
    assert finals == sorted(finals, reverse=True), "jobs sheet not sorted by Final desc"
    # field_ranked sheet has the five decision columns.
    fr = wb["field_ranked"]
    fr_header = [c.value for c in fr[1]]
    assert fr_header == ["career", "field_score", "n_jobs", "n_entry_level", "median_fit"]

    # requirements workbook: canonical-skill frequency table.
    wb2 = load_workbook(paths.requirements_xlsx)
    req = wb2["requirements"]
    req_header = [c.value for c in req[1]]
    assert req_header == ["skill", "required_frequency", "preferred_frequency",
                          "any_frequency", "pct_of_postings", "top_fields"]
    skills = {req.cell(row=r, column=1).value for r in range(2, req.max_row + 1)}
    assert "agents" in skills and "fastapi" in skills

    if paths.scatter_png.exists():
        assert paths.scatter_png.stat().st_size > 0


def test_extract_ignores_diagnostics_and_skips_malformed_raw_cache(
    tmp_dir: Path,
):
    raw_cache = tmp_dir / "raw_cache"
    good = raw_cache / "greenhouse" / "123.json"
    bad = raw_cache / "greenhouse" / "broken.json"
    diagnostic = raw_cache / "_scrapling_failures" / "greenhouse" / "x.json"
    good.parent.mkdir(parents=True)
    diagnostic.parent.mkdir(parents=True)
    write_jsonl(
        good,
        [
            RawPosting(
                board="greenhouse",
                job_id="123",
                url="https://example.test/123",
                fetched_at="2026-08-19T00:00:00Z",
                raw_text="Example role",
            )
        ],
    )
    bad.write_text('{"board":"greenhouse"', encoding="utf-8")
    diagnostic.write_text('{\n  "attempts": []\n}\n', encoding="utf-8")
    selection = tmp_dir / "viable.jsonl"
    write_jsonl(
        selection,
        [JobUrl("greenhouse", "123", "https://example.test/123")],
    )
    messages: list[str] = []
    ctx = SimpleNamespace(
        paths=SimpleNamespace(
            raw_cache=raw_cache,
            processing_job_urls=selection,
            job_urls=selection,
        ),
        log=messages.append,
    )

    records = list(pipeline_run._iter_raw_postings(ctx))

    assert [record.key for record in records] == ["greenhouse:123"]
    assert len(messages) == 1
    assert "skip malformed raw cache file" in messages[0]


# --------------------------------------------------------------------------- #
# Minimal runner: works with or without pytest.
# --------------------------------------------------------------------------- #
def _run_standalone() -> int:
    import tempfile

    tests = [
        test_all_equal_inputs_return_that_value,
        test_one_low_factor_drags_score,
        test_epsilon_prevents_zero_collapse,
        test_monotonic_in_each_axis,
        test_power_mean_ordering,
        test_accessibility_inverts_barrier,
        test_score_row_uses_profile_and_bounds,
        test_final_blend_formula,
        test_field_aggregation_shape,
        test_sensitivity_reports_stability,
    ]
    passed = failed = 0
    for t in tests:
        try:
            t()
            print(f"  PASS  {t.__name__}")
            passed += 1
        except AssertionError as e:
            print(f"  FAIL  {t.__name__}: {e}")
            failed += 1

    # Reporter test needs a temp dir (deletion may be blocked on some mounts, so
    # write into the system temp, not the repo mount).
    with tempfile.TemporaryDirectory() as d:
        try:
            test_reporter_builds_both_workbooks(Path(d))
            print("  PASS  test_reporter_builds_both_workbooks")
            passed += 1
        except AssertionError as e:
            print(f"  FAIL  test_reporter_builds_both_workbooks: {e}")
            failed += 1

    print(f"\n{passed} passed, {failed} failed")
    return 0 if failed == 0 else 1


# pytest fixture for tmp_dir when run under pytest.
try:
    import pytest

    @pytest.fixture
    def tmp_dir(tmp_path):
        return tmp_path
except ImportError:  # pragma: no cover - pytest not installed
    pass


if __name__ == "__main__":
    raise SystemExit(_run_standalone())
